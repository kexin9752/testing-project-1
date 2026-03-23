import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.models.transaction import Transaction


def generate_excel_report(db: Session, user_id: str, report_name: str, report_type: str) -> str:
    """Generate Excel report and return file path"""
    reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "reports")
    os.makedirs(reports_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_name}_{report_type}_{timestamp}.xlsx"
    filepath = os.path.join(reports_dir, filename)

    transactions = db.query(Transaction).filter(Transaction.user_id == user_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["ID", "Type", "Asset", "Amount", "Price", "Status", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.id)
        ws.cell(row=row, column=2, value=t.type.value)
        ws.cell(row=row, column=3, value=t.asset)
        ws.cell(row=row, column=4, value=t.amount)
        ws.cell(row=row, column=5, value=t.price)
        ws.cell(row=row, column=6, value=t.status.value)
        ws.cell(row=row, column=7, value=t.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=len(transactions) + 1, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["G"].width = 25

    wb.save(filepath)
    return filepath